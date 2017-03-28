#!/usr/bin/env python

"""Run epik on all substrates, products and cofactors."""

from sys import argv
import numpy as np
import logging
from openeye import oechem
import openeye
from openmoltools import openeye as omtoe, schrodinger

MAX_ENERGY_PENALTY = 10.0 # kT

def read_molecules(filename):
    """Read a file into an OpenEye molecule (or list of molecules).

    Parameters
    ----------
    filename : str
        The name of the file to read (e.g. mol2, sdf)

    Returns
    -------
    molecule : openeye.oechem.OEMol
        The OEMol molecule read, or a list of molecules if multiple molecules are read.
        If no molecules are read, None is returned.

    """

    ifs = oechem.oemolistream(filename)
    molecules = list()
    for mol in ifs.GetOEMols():
        mol_copy = oechem.OEMol(mol)
        molecules.append(mol_copy)
    ifs.close()

    if len(molecules) == 0:
        return None
    elif len(molecules) == 1:
        return molecules[0]
    else:
        return molecules

def DumpSDData(mol):
    logging.info(("SD data of", mol.GetTitle()))
    #loop over SD data
    for dp in oechem.OEGetSDDataPairs(mol):
        logging.info((dp.GetTag(), ':', dp.GetValue()))
    logging.info('')

def retrieve_url(url, filename):
    import urllib.request, urllib.error, urllib.parse
    # logging.info(url)
    response = urllib.request.urlopen(url)
    html = response.read()
    outfile = open(filename, 'wb')
    outfile.write(html)
    outfile.close()

def read_molecule(filename):
    ifs = oechem.oemolistream()
    ifs.open(filename)
    molecule = oechem.OEMol()
    oechem.OEReadMolecule(ifs, molecule)
    ifs.close()
    return molecule

def fix_mol2_resname(filename, residue_name):
    # Replace <0> substructure names with residue name.
    infile = open(filename, 'r')
    lines = infile.readlines()
    infile.close()
    newlines = [line.replace('<0>', residue_name) for line in lines]
    outfile = open(filename, 'w')
    outfile.writelines(newlines)
    outfile.close()

def write_mol2_preserving_atomnames(filename, molecules, residue_name):
    ofs = oechem.oemolostream()
    ofs.open(filename)
    ofs.SetFlavor(oechem.OEFormat_MOL2, oechem.OEOFlavor_MOL2_GeneralFFFormat)
    try:
        for molecule in molecules:
            oechem.OEWriteMolecule(ofs, molecule)
    except:
        oechem.OEWriteMolecule(ofs, molecules)
    ofs.close()
    fix_mol2_resname(filename, residue_name)

def enumerate_conformations(name, pdbfile=None, smiles=None, pdbname=None, pH=7.4):
    """Run Epik to get protonation states using PDB residue templates for naming.

    Parameters
    ----------
    name : str
       Common name of molecule (used to create subdirectory)
    smiles : str
       Isomeric SMILES string
    pdbname : str
       Three-letter PDB code (e.g. 'DB8')
    """
    # Create output subfolder
    # output_basepath = os.path.join(output_dir, name)
    # if not os.path.isdir(output_basepath):
    #     os.mkdir(output_basepath)
    # output_basepath = os.path.join(output_basepath, name)

    oehandler = openeye.oechem.OEThrow
    # String stream output
    oss = oechem.oeosstream()
    oehandler.SetOutputStream(oss)
    log = "New run:\nPDB code: {pdbname}; Molecule: {name}; pH {pH}\n".format(**locals())
    success_status = True

    if pdbname:
        # Make sure to only use one entry if there are multiple
        if ' ' in pdbname:
            pdbnames = pdbname.split(' ')
            log += "Splitting '%s' into first entry only: '%s'" % (pdbname, pdbnames[0])
            pdbname = pdbnames[0]

        # Retrieve PDB (for atom names)
        url = 'http://ligand-expo.rcsb.org/reports/%s/%s/%s_model.pdb' % (pdbname[0], pdbname, pdbname)
        pdb_filename = name + '-rcsb_download.pdb'
        log += "Retrieving PDB structure from RCSB ligand expo: {}.\n".format(pdb_filename)
        retrieve_url(url, pdb_filename)
        log += "Parsing PDB file.\n"
        pdb_molecule = read_molecule(pdb_filename)

        # Retrieve SDF (for everything else)
        url = 'http://ligand-expo.rcsb.org/reports/%s/%s/%s_model.sdf' % (pdbname[0], pdbname, pdbname)
        sdf_filename = name + '-rcsb_download.sdf'
        log += "Retrieving SDF structure from RCSB ligand expo: {}.\n".format(sdf_filename)
        retrieve_url(url, sdf_filename)
        log += "Parsing SDF file.\n"
        sdf_molecule = read_molecule(sdf_filename)

        # Replace atom names in SDF
        log += "Canonicalizing atom names.\n"
        for (sdf_atom, pdb_atom) in zip(sdf_molecule.GetAtoms(), pdb_molecule.GetAtoms()):
            sdf_atom.SetName(pdb_atom.GetName())
        # Assign Tripos atom types
        log += "Assign atom type names.\n"
        oechem.OETriposAtomTypeNames(sdf_molecule)
        oechem.OETriposBondTypeNames(sdf_molecule)

        oe_molecule = sdf_molecule

        # We already know the residue name
        residue_name = pdbname

    # For the moment, disabling these two types of input
    # elif smiles:
    #     # Generate molecule geometry with OpenEye
    #     logging.info(("Generating molecule {}".format(name)))
    #     oe_molecule = openeye.smiles_to_oemol(smiles)
    #     # Assign Tripos atom types
    #     oechem.OETriposAtomTypeNames(oe_molecule)
    #     oechem.OETriposBondTypeNames(oe_molecule)
    #     try:
    #         logging.info("Charging initial")
    #         write_mol2_preserving_atomnames(name + '-debug.mol2', oe_molecule, 'debug')
    #         oe_molecule = openeye.get_charges(oe_molecule, keep_confs=1)
    #     except RuntimeError as e:
    #         traceback.print_exc()
    #         logging.info(("Skipping molecule " + name))
    #         return
    #     residue_name = re.sub('[^A-Za-z]+', '', name.upper())[:3]
    #     logging.info("resname = %s", residue_name)
    #     oe_molecule.SetTitle(residue_name) # fix iupac name issue with mol2convert
    # elif pdbfile:
    #     residue_name = re.sub('[^A-Za-z]+', '', name.upper())[:3]
    #     logging.info("Loading molecule molecule {0} from {1}".format(name, pdbfile))
    #     oe_molecule = read_molecule(pdbfile)
    #     # Assign Tripos atom types
    #     oechem.OETriposAtomTypeNames(oe_molecule)
    #     oechem.OETriposBondTypeNames(oe_molecule)
    #     try:
    #         logging.info("Charging initial")
    #         write_mol2_preserving_atomnames(name + '-debug.mol2', oe_molecule, 'debug')
    #         oe_molecule = openeye.get_charges(oe_molecule, keep_confs=1)
    #     except RuntimeError as e:
    #         traceback.print_exc()
    #         logging.info(("Skipping molecule " + name))
    #         return
    else:
        raise Exception('Must provide SMILES string or pdbname, or pdbfile')

    # Save mol2 file, preserving atom names
    log += "Running Epik.\n"
    mol2_file_path = name + '-before_epik.mol2'
    write_mol2_preserving_atomnames(mol2_file_path, oe_molecule, residue_name)

    # Run epik on mol2 file
    mae_file_path = name + '-epik.mae'
    schrodinger.run_epik(mol2_file_path, mae_file_path, tautomerize=False, max_structures=50, min_probability=np.exp(-MAX_ENERGY_PENALTY), ph=pH)

    log+= "Epik run completed.\n"
    # Convert maestro file to sdf and mol2
    output_sdf_filename = name + '-after_epik.sdf'
    output_mol2_filename = name + '-after_epik.mol2'
    # logging.info("Creating sdf")
    schrodinger.run_structconvert(mae_file_path, output_sdf_filename)
    # logging.info("Creating mol2")
    schrodinger.run_structconvert(mae_file_path, output_mol2_filename)

    # Read SDF file.
    ifs_sdf = oechem.oemolistream()
    ifs_sdf.SetFormat(oechem.OEFormat_SDF)
    ifs_sdf.open(output_sdf_filename)
    sdf_molecule = oechem.OEGraphMol()

    # Read MOL2 file.
    ifs_mol2 = oechem.oemolistream()
    ifs_mol2.open(output_mol2_filename)
    mol2_molecule = oechem.OEMol()

    # Assign charges.

    # reset count of error handler
    oehandler.Clear()
    log+= "Assigning charges to protonation states.\n"
    charged_molecules = list()
    index = 0
    failed_states = set()
    while oechem.OEReadMolecule(ifs_sdf, sdf_molecule):
        oechem.OEReadMolecule(ifs_mol2, mol2_molecule)

        index += 1
        log += "State {0:d}\n".format(index)
        try:
            # Charge molecule.
            charged_molecule = omtoe.get_charges(mol2_molecule, max_confs=100, strictStereo=False, normalize=True, keep_confs=None)

            # If more than one warning was raised
            # This still allows the single line "trans conformer warning"
            if oehandler.Count(0) > 1:
                raise RuntimeError(str(oss))

            else:
                oss.clear()

            oehandler.Clear()

            log += "Charging completed.\n"
            # Assign Tripos types
            oechem.OETriposAtomTypeNames(charged_molecule)
            oechem.OETriposBondTypeNames(charged_molecule)
            # Store tags.
            oechem.OECopySDData(charged_molecule, sdf_molecule)
            # Store molecule
            charged_molecules.append(charged_molecule)
        except Exception as e:
            failed_states.add(index)
            logging.info(e)
            log += "Skipping state because of failed charging.\n"
            log += str(e)
            log += "\n"
            success_status = False

    # Clean up
    ifs_sdf.close()
    ifs_mol2.close()

    # Write state penalties.
    outfile = open(name + '-state-penalties.out', 'w')
    for (index, charged_molecule) in enumerate(charged_molecules):
        # Get Epik data.
        log += "Writing Epik data for state {:d}\n".format(index + 1)
        epik_Ionization_Penalty = float(oechem.OEGetSDData(charged_molecule, "r_epik_Ionization_Penalty"))
        epik_Ionization_Penalty_Charging = float(oechem.OEGetSDData(charged_molecule, "r_epik_Ionization_Penalty_Charging"))
        epik_Ionization_Penalty_Neutral = float(oechem.OEGetSDData(charged_molecule, "r_epik_Ionization_Penalty_Neutral"))
        epik_State_Penalty = float(oechem.OEGetSDData(charged_molecule, "r_epik_State_Penalty"))
        epik_Tot_Q = int(oechem.OEGetSDData(charged_molecule, "i_epik_Tot_Q"))

        outfile.write('%16.8f\n' % epik_State_Penalty)
    outfile.close()

    # Write as PDB
    charged_pdb_filename = name + '-charged_output.pdb'
    failed_pdb_filename = name + '-failures.pdb'
    ofs = oechem.oemolostream(charged_pdb_filename)
    ofs_fail = oechem.oemolostream(failed_pdb_filename)
    flavor = oechem.OEOFlavor_PDB_CurrentResidues | oechem.OEOFlavor_PDB_ELEMENT | oechem.OEOFlavor_PDB_BONDS | oechem.OEOFlavor_PDB_HETBONDS | oechem.OEOFlavor_PDB_BOTH
    ofs.SetFlavor(oechem.OEFormat_PDB, flavor)
    ofs_fail.SetFlavor(oechem.OEFormat_PDB,flavor)
    for (index, charged_molecule) in enumerate(charged_molecules):
        # Fix residue names
        for atom in charged_molecule.GetAtoms():
            residue = oechem.OEAtomGetResidue(atom)
            residue.SetName(residue_name)
            oechem.OEAtomSetResidue(atom, residue)
        oechem.OEWriteMolecule(ofs, charged_molecule)
        if index + 1 in failed_states:
            oechem.OEWriteMolecule(ofs_fail, charged_molecule)
    ofs.close()
    ofs_fail.close()

    # Write molecules as mol2.
    charged_mol2_filename = name + '-charged_output.mol2'
    write_mol2_preserving_atomnames(charged_mol2_filename, charged_molecules, residue_name)
    log += "Run completed.\n"
    if success_status:
        log += "Status: Success\n"
    else:
        log += "Status: Failure\n"

    with open("log.txt", 'w') as logfile:
        logfile.write(log)

    return log, success_status


def main(args):

    import os
    from openpyxl import load_workbook
    import shutil
    from openpyxl.comments import Comment
    workbook = load_workbook('Queue.xlsx')
    to_run = workbook['To Run']
    successes = workbook['Successes']
    failures = workbook['Failures']
    topdir = os.getcwd()
    # Makes the OE license available globally
    os.environ['OE_LICENSE'] = os.getcwd() + '\\oe_license.txt'
    os.environ['SCHRODINGER'] = 'C:\\Program Files\\Schrodinger2015-3'
    # Create the directories to store output
    directories = ['InProgress', 'Failures', 'Successes']

    if not os.path.isdir('Results'):
        os.mkdir('Results')
    os.chdir('Results')
    if not os.path.isdir('Failures'):
        os.mkdir('Failures')
    if not os.path.isdir('Successes'):
        os.mkdir('Successes')
    if not os.path.isdir('InProgress'):
        os.mkdir('InProgress')
    # Skip header
    for row_in_sheet in to_run.iter_rows(min_row=2):

        # Skip empty ligand
        if row_in_sheet[1].value is None:
            continue

        ligand_dirname = row_in_sheet[2].value

        # Make sure the ligand directory exists in all subdirectories
        for directory in directories:
            datadir = os.path.abspath('{directory}/{ligand_dirname}'.format(**locals()))
            if not os.path.isdir(datadir):
                os.mkdir(datadir)

        os.chdir("InProgress")
        progress_dir = os.getcwd()
        os.chdir(ligand_dirname)
        ligand_absdir = os.getcwd()

        # Parse the pH values provided in the XLS sheet
        phs = list()
        ph_list_as_string = str(row_in_sheet[3].value)
        for entry in ph_list_as_string.split(sep=","):
            phs.append(float(entry))

        # Run the procedure for each requested pH
        for ph in phs:
            pH_dir = 'pH_{}'.format(ph)
            ph_faildir = os.path.abspath("{progress_dir}/../Failures/{ligand_dirname}/{pH_dir}".format(**locals()))
            ph_successdir = os.path.abspath("{progress_dir}/../Successes/{ligand_dirname}/{pH_dir}".format(**locals()))
            if os.path.isdir(ph_faildir):
                print("Please delete {} if you wish to retry this ligand.".format(ph_faildir))
                continue
            elif os.path.isdir(ph_successdir):
                print("Already completed {}".format(ph_successdir))
                continue

            logger = logging.getLogger()
            logger.setLevel(logging.INFO)

            # Create output directory, delete old in process data if present
            if not os.path.isdir(pH_dir):
                os.mkdir(pH_dir)
            else:
                shutil.rmtree(pH_dir)
                os.mkdir(pH_dir)
                continue
            os.chdir(pH_dir)

            log, success = enumerate_conformations(ligand_dirname + '_' + pH_dir, pdbname=row_in_sheet[0].value, pH=ph)
            os.chdir(ligand_absdir)

            comment = Comment(log, __file__)

            # Update excel sheets and directories
            if success:
                os.rename(os.path.abspath(pH_dir), ph_successdir)
                row_ = successes.max_row + 1
                for col, cell in enumerate(row_in_sheet, start=1):
                    if col < 4:
                        newcell = successes.cell(row=row_, column=col, value=cell.value)
                    elif col == 4:
                        newcell = successes.cell(row=row_, column=col, value=ph)

                    newcell = successes.cell(row=row_, column=5, value="See comment for log.")
                    newcell.comment = comment
            else:
                os.rename(os.path.abspath(pH_dir), ph_faildir)
                row_ = failures.max_row + 1
                for col, cell in enumerate(row_in_sheet, start=1):
                    if col < 4:
                        newcell = failures.cell(row=row_, column=col, value=cell.value)
                    elif col == 4:
                        newcell = failures.cell(row=row_, column=col, value=ph)

                    newcell = failures.cell(row=row_, column=5, value="See comment for log.")
                    newcell.comment = comment

        os.chdir(os.path.abspath("{topdir}/Results".format(**locals())))
        # Clean up sheet by deleting the completed run
        for cell in row_in_sheet:
            cell.value = ""

    os.chdir(topdir)
    workbook.save('Queue.xlsx')


if __name__ == '__main__':
    main(args=argv)